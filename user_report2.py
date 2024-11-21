import tkinter as tk
from tkinter import ttk, messagebox
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import shutil
import configparser
import csv
from datetime import datetime

class ADReportGUI:
    def __init__(self, master):
        self.master = master
        master.title("Active Directory OU Bazlı Rapor Oluşturucu @by Önder AKÖZ")
        master.geometry("500x400")

        # Yapılandırma dosyasını oku
        self.config = configparser.ConfigParser()
        self.config.read('config.ini', encoding='utf-8')

        # GUI değişkenlerini oluştur
        self.month_range = tk.StringVar(value="3")
        self.recipient_email = tk.StringVar()
        self.excel_password = tk.StringVar()
        self.include_sub_ous = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar()

        # GUI bileşenlerini oluştur
        self.create_widgets()
        
    def update_status(self, message, error=False):
        """Durum çubuğunu güncelle"""
        self.status_var.set(message)
        self.status_bar.configure(foreground="red" if error else "black")
        self.master.update_idletasks()
    def create_widgets(self):
        # Ana Frame
        main_frame = ttk.Frame(self.master)
        main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Rapor Ayarları Grubu
        settings_frame = ttk.LabelFrame(main_frame, text="Rapor Ayarları")
        settings_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        # Ay Sayısı
        ttk.Label(settings_frame, text="Rapor için ay sayısı:").grid(row=0, column=0, padx=5, pady=5)
        self.month_range_entry = ttk.Entry(settings_frame, textvariable=self.month_range)
        self.month_range_entry.grid(row=0, column=1, padx=5, pady=5)

        # E-posta
        ttk.Label(settings_frame, text="Alıcı E-posta:").grid(row=1, column=0, padx=5, pady=5)
        self.recipient_email_entry = ttk.Entry(settings_frame, textvariable=self.recipient_email)
        self.recipient_email_entry.grid(row=1, column=1, padx=5, pady=5)

        # Excel Şifresi
        ttk.Label(settings_frame, text="Excel Şifresi:").grid(row=2, column=0, padx=5, pady=5)
        self.excel_password_entry = ttk.Entry(settings_frame, show="*", textvariable=self.excel_password)
        self.excel_password_entry.grid(row=2, column=1, padx=5, pady=5)

        # OU Seçim Grubu
        ou_frame = ttk.LabelFrame(main_frame, text="OU Seçenekleri")
        ou_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        # Alt OU Seçeneği
        ttk.Checkbutton(
            ou_frame,
            text="Alt OU'ları dahil et",
            variable=self.include_sub_ous
        ).grid(row=0, column=0, padx=5, pady=5)

        # Durum Çubuğu
        self.status_bar = ttk.Label(
            main_frame,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor="w"
        )
        self.status_bar.grid(row=3, column=0, sticky="ew", padx=5)

        # Butonlar
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, padx=5, pady=5)

        self.test_mail_button = ttk.Button(
            button_frame,
            text="Test Mail Gönder",
            command=self.send_test_mail
        )
        self.test_mail_button.grid(row=0, column=0, padx=5, pady=5)

        self.generate_button = ttk.Button(
            button_frame,
            text="Rapor Oluştur ve Gönder",
            command=self.generate_report
        )
        self.generate_button.grid(row=0, column=1, padx=5, pady=5)

        # Grid Yapılandırması
        self.master.columnconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        settings_frame.columnconfigure(1, weight=1)
        ou_frame.columnconfigure(0, weight=1)
    def send_test_mail(self):
        """Test e-postası gönder"""
        try:
            self.update_status("Test maili gönderiliyor...")
            
            msg = MIMEMultipart()
            msg['From'] = self.config['SMTP']['sender_email']
            msg['To'] = self.recipient_email.get()
            msg['Subject'] = "Test Mail - AD OU Rapor Oluşturucu"

            body = """
            Bu, AD OU Rapor Oluşturucu uygulamasından gönderilen bir test e-postasıdır.
            
            E-posta ayarlarınız doğru çalışıyor.
            """
            msg.attach(MIMEText(body, 'plain', 'utf-8'))

            self.send_email(msg)
            messagebox.showinfo("Başarılı", "Test maili başarıyla gönderildi!")
            self.update_status("Test maili gönderildi")
        
        except Exception as e:
            error_msg = f"Test maili gönderilirken bir hata oluştu: {str(e)}"
            messagebox.showerror("Hata", error_msg)
            self.update_status(error_msg, error=True)

    def send_email(self, msg):
        """E-posta gönderme işlemi"""
        with smtplib.SMTP(self.config['SMTP']['server'], self.config['SMTP']['port']) as server:
            server.starttls()
            server.login(self.config['SMTP']['sender_email'], self.config['SMTP']['password'])
            server.send_message(msg)

    def send_report_email(self, month_range):
        """Raporu e-posta ile gönder"""
        msg = MIMEMultipart()
        msg['From'] = self.config['SMTP']['sender_email']
        msg['To'] = self.recipient_email.get()
        msg['Subject'] = f"Active Directory OU Bazlı Rapor - Son {month_range} Ay"

        body = f"""
        Merhaba,

        Ekte son {month_range} aya ait Active Directory OU bazlı raporu bulunmaktadır.
        
        Rapor İçeriği:
        - Her Organizational Unit (OU) için ayrı sayfa
        - Domain Adminler listesi
        - Kullanıcı detayları (Ad, E-posta, Unvan, Departman, Yönetici, Son Oturum, Gruplar)
        
        Not: Excel dosyası şifrelidir. Şifreyi ayrıca alınız.
        
        İyi çalışmalar.
        """
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # Excel dosyasını ekle
        excel_file = f"AD_OU_Raporu_Son_{month_range}_Ay.xlsx"
        with open(excel_file, 'rb') as file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="{excel_file}"'
            )
            msg.attach(part)

        self.send_email(msg)
    def create_excel_report(self, month_range, excel_password):
        """Excel raporu oluştur"""
        try:
            temp_dir = self.run_powershell_script(month_range)
            
            wb = openpyxl.Workbook()
            
            # İlk sayfa - Domain Adminler
            ws_admins = wb.active
            ws_admins.title = "Domain Adminler"
            admin_csv = os.path.join(temp_dir, "domain_admins.csv")
            if os.path.exists(admin_csv):
                self.populate_worksheet(ws_admins, admin_csv)
            
            # OU bazlı sayfalar
            for file in os.listdir(temp_dir):
                if file.startswith("users_") and file.endswith(".csv"):
                    ou_name = file[6:-4].replace("_", " ")
                    ws = wb.create_sheet(ou_name)
                    self.populate_worksheet(ws, os.path.join(temp_dir, file))
            
            # Excel'i kaydet ve şifrele
            excel_file = f"AD_OU_Raporu_Son_{month_range}_Ay.xlsx"
            wb.save(excel_file)
            self.encrypt_excel_file(excel_file, excel_password)
            
        finally:
            # Geçici dosyaları temizle
            self.cleanup_files()

    def populate_worksheet(self, ws, csv_file):
        """Excel sayfasını CSV verileriyle doldur"""
        # Stil tanımlamaları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # CSV'yi oku ve verileri aktar
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader)
            
            # Başlıkları yaz ve formatla
            for col, title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Verileri yaz ve formatla
            for row_idx, row in enumerate(reader, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Son oturum tarihi için özel format
                    if headers[col_idx-1] == "LastLogonDate" and value:
                        try:
                            if isinstance(value, str):
                                date_value = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                                cell.value = date_value
                                cell.number_format = "dd.mm.yyyy hh:mm"
                        except ValueError:
                            pass
                    
                    # Enabled durumu için özel format
                    if headers[col_idx-1] == "Enabled":
                        cell.alignment = Alignment(horizontal="center")
                        if value.lower() == "true":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value.lower() == "false":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Maximum 50 karakter
        
        # Başlık satırını dondur
        ws.freeze_panes = "A2"
    def encrypt_excel_file(self, excel_file, password):
        """Excel dosyasını şifrele"""
        temp_file = "temp_" + excel_file
        
        try:
            # PowerShell komutu oluştur
            ps_script = f"""
            $filePath = "{os.path.abspath(excel_file)}"
            $tempPath = "{os.path.abspath(temp_file)}"
            
            # Excel uygulamasını başlat
            $excel = New-Object -ComObject Excel.Application
            $excel.Visible = $false
            $excel.DisplayAlerts = $false
            
            Write-Host "Excel açılıyor..."
            
            try {{
                # Excel dosyasını aç
                $workbook = $excel.Workbooks.Open($filePath)
                Write-Host "Dosya açıldı: $filePath"
                
                # Dosyayı şifreli olarak kaydet
                $workbook.Password = "{password}"
                $workbook.SaveAs($tempPath, 51, "{password}")  # 51: xlOpenXMLWorkbook
                Write-Host "Dosya şifrelendi ve kaydedildi: $tempPath"
                
                # Dosyayı kapat
                $workbook.Close($false)
                [System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null
                
            }} catch {{
                Write-Error "Hata oluştu: $($_.Exception.Message)"
                throw
            }} finally {{
                # Excel'i kapat ve kaynakları temizle
                $excel.Quit()
                [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
                [System.GC]::Collect()
                [System.GC]::WaitForPendingFinalizers()
            }}
            """
            
            # PowerShell'i çalıştır
            process = subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
                capture_output=True,
                text=True,
                encoding='utf-8'
            )
            
            # Hata kontrolü
            if process.returncode != 0:
                raise Exception(f"PowerShell hatası: {process.stderr}")
            
            # Başarılı olursa dosyaları değiştir
            if os.path.exists(temp_file):
                if os.path.exists(excel_file):
                    os.remove(excel_file)
                os.rename(temp_file, excel_file)
            else:
                raise Exception("Şifrelenmiş dosya oluşturulamadı")
            
        except Exception as e:
            # Hata durumunda geçici dosyayı temizle
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
            # Hata mesajını yukarı ilet
            raise Exception(f"Excel şifreleme hatası: {str(e)}")
    def cleanup_files(self):
        """Geçici dosyaları temizle"""
        temp_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'ADReports')
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"Temizlik sırasında hata: {str(e)}")
    def run_powershell_script(self, month_range):
        """PowerShell scripti çalıştır ve AD verilerini al"""
        # Geçici klasör oluştur
        temp_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'ADReports')
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)

        # Search scope'u belirle
        search_scope = "Subtree" if self.include_sub_ous.get() else "OneLevel"

        ps_script = f"""
        # Parametreler ve encoding ayarları
        $ErrorActionPreference = "Stop"
        $tempDir = '{temp_dir.replace('\\', '/')}'
        $months = {month_range}
        $searchScope = '{search_scope}'
        $lastLogonDate = (Get-Date).AddMonths(-$months)
        
        [Console]::OutputEncoding = [System.Text.Encoding]::UTF8
        $OutputEncoding = [Console]::OutputEncoding

        Write-Host "Başlangıç parametreleri:"
        Write-Host "Ay sayısı: $months"
        Write-Host "Son oturum tarihi: $lastLogonDate"
        Write-Host "Arama kapsamı: $searchScope"
        Write-Host "Geçici klasör: $tempDir"

        # Domain Adminleri raporla
        try {{
            Write-Host "`nDomain Admin raporu hazırlanıyor..."
            
            $admins = Get-ADGroupMember "Domain Admins" -ErrorAction Stop |
                     Get-ADUser -Properties LastLogonDate,MemberOf,Enabled,Department,Title,Manager,EmailAddress -ErrorAction Stop
            
            if ($admins) {{
                Write-Host "Domain Admin sayısı: $($admins.Count)"
                $adminReport = @()
                
                foreach ($admin in $admins) {{
                    Write-Host "İşleniyor: $($admin.Name)" -ForegroundColor Gray
                    
                    $manager = "Yok"
                    if ($admin.Manager) {{
                        try {{
                            $manager = (Get-ADUser $admin.Manager -ErrorAction Stop).Name
                        }} catch {{
                            $manager = "Bulunamadı"
                        }}
                    }}
                    
                    $groups = "Yok"
                    if ($admin.MemberOf) {{
                        try {{
                            $groups = ($admin.MemberOf | 
                                     ForEach-Object {{ (Get-ADGroup $_ -ErrorAction Stop).Name }} | 
                                     Where-Object {{ $_ }} | 
                                     Sort-Object) -join '; '
                        }} catch {{
                            $groups = "Alınamadı"
                        }}
                    }}
                    
                    $adminReport += [PSCustomObject]@{{
                        Name = $admin.Name
                        SamAccountName = $admin.SamAccountName
                        EmailAddress = $admin.EmailAddress
                        Title = $admin.Title
                        Department = $admin.Department
                        Manager = $manager
                        LastLogonDate = $admin.LastLogonDate
                        Groups = $groups
                        Enabled = $admin.Enabled
                    }}
                }}
                
                $adminCsvPath = Join-Path $tempDir "domain_admins.csv"
                $adminReport | Export-Csv -Path $adminCsvPath -NoTypeInformation -Encoding UTF8 -Force
                Write-Host "Domain Admin raporu kaydedildi: $adminCsvPath" -ForegroundColor Green
            }}
        }} catch {{
            Write-Warning "Domain Admin raporu hazırlanırken hata: $($_.Exception.Message)"
        }}

        # OU bazlı rapor
        try {{
            Write-Host "`nOU raporu hazırlanıyor..."
            $OUs = Get-ADOrganizationalUnit -Filter * -ErrorAction Stop
            Write-Host "Toplam OU sayısı: $($OUs.Count)"
            
            foreach ($OU in $OUs) {{
                try {{
                    Write-Host "`nOU işleniyor: $($OU.Name)"
                    
                    $users = Get-ADUser -SearchBase $OU.DistinguishedName `
                                      -SearchScope $searchScope `
                                      -Filter {{LastLogonDate -ge $lastLogonDate}} `
                                      -Properties LastLogonDate,MemberOf,Enabled,Department,Title,Manager,EmailAddress `
                                      -ErrorAction Stop
                    
                    if ($users) {{
                        Write-Host "Bulunan kullanıcı sayısı: $($users.Count)"
                        $userReport = @()
                        
                        foreach ($user in $users) {{
                            Write-Host "Kullanıcı işleniyor: $($user.Name)" -ForegroundColor Gray
                            
                            $manager = "Yok"
                            if ($user.Manager) {{
                                try {{
                                    $manager = (Get-ADUser $user.Manager -ErrorAction Stop).Name
                                }} catch {{
                                    $manager = "Bulunamadı"
                                }}
                            }}
                            
                            $groups = "Yok"
                            if ($user.MemberOf) {{
                                try {{
                                    $groups = ($user.MemberOf | 
                                             ForEach-Object {{ (Get-ADGroup $_ -ErrorAction Stop).Name }} | 
                                             Where-Object {{ $_ }} | 
                                             Sort-Object) -join '; '
                                }} catch {{
                                    $groups = "Alınamadı"
                                }}
                            }}
                            
                            $userReport += [PSCustomObject]@{{
                                OU = $OU.Name
                                Name = $user.Name
                                SamAccountName = $user.SamAccountName
                                EmailAddress = $user.EmailAddress
                                Title = $user.Title
                                Department = $user.Department
                                Manager = $manager
                                LastLogonDate = $user.LastLogonDate
                                Groups = $groups
                                Enabled = $user.Enabled
                            }}
                        }}
                        
                        $csvPath = Join-Path $tempDir "users_$($OU.Name -replace '[^\w\-]', '_').csv"
                        $userReport | Export-Csv -Path $csvPath -NoTypeInformation -Encoding UTF8 -Force
                        Write-Host "OU raporu kaydedildi: $csvPath" -ForegroundColor Green
                    }} else {{
                        Write-Host "Bu OU'da kullanıcı bulunamadı" -ForegroundColor Yellow
                    }}
                }} catch {{
                    Write-Warning "OU işlenirken hata: $($_.Exception.Message)"
                    continue
                }}
            }}
        }} catch {{
            Write-Warning "OU raporu hazırlanırken hata: $($_.Exception.Message)"
        }}
        """

        try:
            # PowerShell'i çalıştır
            self.update_status("PowerShell scripti çalıştırılıyor...")
            process = subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
                capture_output=True,
                text=True,
                encoding='utf-8'
            )

            # Hata kontrolü
            if process.returncode != 0:
                error_msg = f"PowerShell hatası:\n{process.stderr}"
                print(error_msg)
                raise Exception(error_msg)

            # İşlem durumunu yazdır
            print("\nPowerShell Çıktısı:")
            print(process.stdout)
            
            if process.stderr:
                print("\nPowerShell Hataları:")
                print(process.stderr)

            self.update_status("Veri toplama tamamlandı")
            return temp_dir

        except Exception as e:
            self.update_status(f"Hata: {str(e)}", error=True)
            if os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                except:
                    pass
            raise e
    def generate_report(self):
        """Ana rapor oluşturma metodu"""
        if not self.excel_password.get():
            messagebox.showerror("Hata", "Lütfen Excel dosyası için bir şifre girin.")
            return

        if not self.recipient_email.get():
            messagebox.showerror("Hata", "Lütfen alıcı e-posta adresini girin.")
            return

        try:
            # Butonları devre dışı bırak
            self.test_mail_button.state(['disabled'])
            self.generate_button.state(['disabled'])
            
            # Rapor oluştur
            self.update_status("Rapor oluşturuluyor...")
            self.create_excel_report(self.month_range.get(), self.excel_password.get())
            
            # E-posta gönder
            self.update_status("Rapor gönderiliyor...")
            self.send_report_email(self.month_range.get())
            
            # Başarı mesajı
            messagebox.showinfo("Başarılı", "OU bazlı rapor oluşturuldu ve başarıyla gönderildi!")
            self.update_status("İşlem tamamlandı")

        except Exception as e:
            error_msg = f"Rapor oluşturulurken bir hata oluştu: {str(e)}"
            messagebox.showerror("Hata", error_msg)
            self.update_status(error_msg, error=True)
        
        finally:
            # Butonları tekrar aktif et
            self.test_mail_button.state(['!disabled'])
            self.generate_button.state(['!disabled'])
            self.cleanup_files()


if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = ADReportGUI(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("Kritik Hata", f"Uygulama başlatılırken bir hata oluştu:\n{str(e)}")                
                    
